import json
from openpyxl import load_workbook
import re

dataset = json.load(open("resources/LTAMRTStationExitGEOJSON.geojson"))
wb = load_workbook(filename="singapore_stations.xlsx", read_only=False)
sheet = wb["stations"]

already_found = set()
aggregate = []
for feature in dataset["features"]:
    raw_name = feature["properties"]["STATION_NA"]
    if "LRT STATION" in raw_name or raw_name in already_found:
        continue
    already_found.add(raw_name)

    name = re.sub(r"\sMRT\sSTATION", "", raw_name)
    lon = feature["geometry"]["coordinates"][0]
    lat = feature["geometry"]["coordinates"][1]
    for row in range(1, 150):
        cell_value = str(sheet.cell(row, 2).value)
        if name in str.upper(cell_value):
            sheet[f"F{row}"] = lon
            sheet[f"G{row}"] = lat
            print(cell_value)
    
    wb.save(filename="test.xlsx")
    aggregate.append({'name': name, 'lon': lon, 'lat': lat})

print(aggregate)
